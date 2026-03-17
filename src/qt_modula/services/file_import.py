"""Typed local file import services."""

from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import orjson
from openpyxl import load_workbook  # type: ignore[import-untyped]

from qt_modula.services.errors import ServiceError

_PROVIDER = "file_import"
_TABLE_FORMATS = {"csv", "jsonl", "xlsx"}
_TABLE_EXTENSION_MAP = {
    ".csv": "csv",
    ".jsonl": "jsonl",
    ".xlsx": "xlsx",
}


@dataclass(frozen=True, slots=True)
class TextImportRequest:
    """Typed text import request."""

    path: Path


@dataclass(frozen=True, slots=True)
class TextImportResult:
    """Typed text import response."""

    path: Path
    content: str
    char_count: int
    line_count: int


@dataclass(frozen=True, slots=True)
class JsonImportRequest:
    """Typed JSON import request."""

    path: Path


@dataclass(frozen=True, slots=True)
class JsonImportResult:
    """Typed JSON import response."""

    path: Path
    document: Any
    keys: list[str]
    item_count: int


@dataclass(frozen=True, slots=True)
class TableImportRequest:
    """Typed tabular import request."""

    path: Path
    format: str = "auto"
    sheet_name: str = ""


@dataclass(frozen=True, slots=True)
class TableImportResult:
    """Typed tabular import response."""

    path: Path
    rows: list[dict[str, Any]]
    row_count: int
    column_count: int
    columns: list[str]
    format: str
    sheet_name: str = ""


def read_text_file(request: TextImportRequest) -> TextImportResult:
    """Read one UTF-8 or UTF-8-SIG text file."""

    path = _validated_file_path(request.path)
    payload = path.read_bytes()
    try:
        content = payload.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ServiceError(
            kind="validation",
            message=f"File is not valid UTF-8 text: {path}.",
            provider=_PROVIDER,
        ) from exc
    return TextImportResult(
        path=path,
        content=content,
        char_count=len(content),
        line_count=_line_count(content),
    )


def read_json_file(request: JsonImportRequest) -> JsonImportResult:
    """Read one JSON file with `orjson`."""

    path = _validated_file_path(request.path)
    if path.suffix.lower() != ".json":
        raise ServiceError(
            kind="validation",
            message=f"Unsupported JSON file extension: {path.suffix or '<none>'}.",
            provider=_PROVIDER,
        )
    try:
        document = orjson.loads(path.read_bytes())
    except orjson.JSONDecodeError as exc:
        raise ServiceError(
            kind="validation",
            message=f"Invalid JSON file: {exc}.",
            provider=_PROVIDER,
        ) from exc

    if isinstance(document, dict):
        keys = sorted(str(key) for key in document)
        item_count = len(document)
    elif isinstance(document, list):
        keys = []
        item_count = len(document)
    else:
        keys = []
        item_count = 1

    return JsonImportResult(
        path=path,
        document=document,
        keys=keys,
        item_count=item_count,
    )


def read_table_file(request: TableImportRequest) -> TableImportResult:
    """Read one CSV, JSONL, or XLSX table."""

    path = _validated_file_path(request.path)
    format_token = _normalized_table_format(request.format, path=path)
    sheet_name = request.sheet_name.strip()

    if format_token == "csv":
        rows, columns = _read_csv_rows(path)
        resolved_sheet_name = ""
    elif format_token == "jsonl":
        rows, columns = _read_jsonl_rows(path)
        resolved_sheet_name = ""
    else:
        rows, columns, resolved_sheet_name = _read_xlsx_rows(path, sheet_name=sheet_name)

    return TableImportResult(
        path=path,
        rows=rows,
        row_count=len(rows),
        column_count=len(columns),
        columns=columns,
        format=format_token,
        sheet_name=resolved_sheet_name,
    )


def _validated_file_path(value: Path | str) -> Path:
    path = Path(value).expanduser().resolve(strict=False)
    if not path.exists():
        raise ServiceError(
            kind="not_found",
            message=f"File not found: {path}.",
            provider=_PROVIDER,
        )
    if not path.is_file():
        raise ServiceError(
            kind="validation",
            message=f"Import path must point to a file: {path}.",
            provider=_PROVIDER,
        )
    return path


def _line_count(text: str) -> int:
    return len(text.splitlines()) if text else 0


def _normalized_table_format(value: str, *, path: Path) -> str:
    token = value.strip().lower().lstrip(".")
    if token in {"", "auto"}:
        inferred = _TABLE_EXTENSION_MAP.get(path.suffix.lower())
        if inferred is None:
            raise ServiceError(
                kind="validation",
                message=(
                    "Could not infer a table format from the file extension. "
                    "Use .csv, .jsonl, .xlsx, or set the format explicitly."
                ),
                provider=_PROVIDER,
            )
        return inferred
    if token in _TABLE_FORMATS:
        return token
    raise ServiceError(
        kind="validation",
        message=f"Unsupported table import format '{value}'.",
        provider=_PROVIDER,
    )


def _unique_headers(raw_headers: tuple[Any, ...]) -> list[str]:
    result: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(raw_headers):
        token = str(value).strip() if value is not None else ""
        if not token:
            token = f"c{index}"
        base = token
        suffix = seen.get(base, 0)
        while token in seen:
            suffix += 1
            token = f"{base}_{suffix}"
        seen[base] = suffix
        seen[token] = 0
        result.append(token)
    return result


def _ordered_columns(rows: list[dict[str, Any]]) -> list[str]:
    columns: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for key in row:
            token = str(key)
            if token in seen:
                continue
            seen.add(token)
            columns.append(token)
    return columns


def _read_csv_rows(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        header_row = next(reader, None)
        if header_row is None:
            return [], []
        columns = _unique_headers(tuple(header_row))
        rows: list[dict[str, Any]] = []
        for raw_row in reader:
            row = {
                columns[index]: raw_row[index] if index < len(raw_row) else ""
                for index in range(len(columns))
            }
            if any(value not in {"", None} for value in row.values()):
                rows.append(row)
        return rows, columns


def _read_jsonl_rows(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    rows: list[dict[str, Any]] = []
    with path.open("rb") as handle:
        for line_number, raw_line in enumerate(handle, start=1):
            payload = raw_line.strip()
            if not payload:
                continue
            try:
                loaded = orjson.loads(payload)
            except orjson.JSONDecodeError as exc:
                raise ServiceError(
                    kind="validation",
                    message=f"Invalid JSONL row {line_number}: {exc}.",
                    provider=_PROVIDER,
                ) from exc
            if isinstance(loaded, dict):
                row = {str(key): value for key, value in loaded.items()}
            else:
                row = {"value": loaded}
            rows.append(row)
    return rows, _ordered_columns(rows)


def _read_xlsx_rows(path: Path, *, sheet_name: str) -> tuple[list[dict[str, Any]], list[str], str]:
    workbook = load_workbook(filename=str(path), data_only=True, read_only=True)
    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ServiceError(
                    kind="validation",
                    message=f"Sheet '{sheet_name}' was not found in {path.name}.",
                    provider=_PROVIDER,
                )
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active

        iterator = sheet.iter_rows(values_only=True)
        header_row = next(iterator, None)
        if header_row is None:
            return [], [], str(sheet.title)
        columns = _unique_headers(tuple(header_row))
        rows: list[dict[str, Any]] = []
        for raw_row in iterator:
            row = {
                columns[index]: raw_row[index] if index < len(raw_row) else None
                for index in range(len(columns))
            }
            if any(value not in {"", None} for value in row.values()):
                rows.append(row)
        return rows, columns, str(sheet.title)
    finally:
        workbook.close()
